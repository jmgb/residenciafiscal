# residenciafiscal.py
# Recorre la carpeta de sentencias PDFs, extrae texto (con marcadores de página),
# llama a un LLM con tu prompt de sistema (prompt.py), y genera:
#  - output.jsonl (una línea JSON por sentencia PDF)
#  - output.csv (una fila por sentencia PDF, con datos estructurados)
#
# Usa la función gpt_request() universal que soporta múltiples proveedores
# y reasoning_effort para modelos GPT-5+.
#
# Requisitos:
#   make setup          (uv crea .venv con Python 3.13 e instala desde uv.lock)
#
# Uso básico:
#   make run            (equivale a: uv run python src/residenciafiscal.py)
#   make run-sample     (1 solo PDF, prueba rápida)
#
# Uso con argumentos personalizados:
#   make run MODEL=gpt-4 INPUT=/ruta/a/pdfs OUTPUT=/ruta/salida
#   uv run python src/residenciafiscal.py --help
#
# Este mismo pipeline se expone por HTTP en src/api/main.py (`make dev-api`,
# también incluido en `make dev`).

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from dotenv import load_dotenv
from pypdf import PdfReader

from config import (
    ALLOWED_KEYS,
    ARGUMENT_HELP,
    BATCH_SIZE,
    CSV_COLUMN_ORDER,
    DEFAULT_CSV_NAME,
    DEFAULT_INPUT_DIR,
    DEFAULT_JSONL_NAME,
    DEFAULT_MAX_FILES,
    DEFAULT_MISSING_VALUE,
    DEFAULT_MODEL,
    DEFAULT_OUTPUT_DIR,
    ENFORCE_ALLOWED_KEYS,
    KEY_SENTENCIAS_FILE,
    PAGE_MARKER_FMT,
    REASONING_EFFORT,
    REQUIRED_FIELDS,
    SCRIPT_DESCRIPTION,
    SENTENCIA_CLAVE_MODEL,
    VALID_CATEGORIAS_PRUEBA,
    # Enums para validación de schema
    VALID_CRITERIOS,
    VALID_RESULTADO_FINAL,
    VALID_TIEBREAKER_PASOS,
    detect_provider,
)

# Intenta importar gpt_request de ai_service_adapter
try:
    from ai_service_adapter import gpt_request_for_sentencia

    USE_GPT_REQUEST = True
except ImportError:
    USE_GPT_REQUEST = False

# Importa tu prompt
try:
    from prompt import system_prompt as SYSTEM_PROMPT  # type: ignore
except Exception as e:
    raise RuntimeError(
        "No pude importar system_prompt desde prompt.py. "
        "Asegúrate de tener system_prompt = '''...''' en prompt.py."
    ) from e


# ============================================================================
# LOGGING
# ============================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    force=True,
)
logger = logging.getLogger(__name__)


# ============================================================================
# UTILIDADES
# ============================================================================


def get_timestamp_suffix() -> str:
    """Genera un sufijo con fecha y hora en formato DDMMYYYY_HHMMSS.

    Returns:
        str: Sufijo timestamp (e.g., "21122026_120012")
    """
    now = datetime.now()
    return now.strftime("%d%m%Y_%H%M%S")


# ============================================================================
# CLIENT INITIALIZATION
# ============================================================================


def initialize_client(ai_model: str) -> str:
    """Initialize the AI client needed for the selected model.

    Detects the provider from the model name and validates required API keys
    are available before processing starts. Fails fast if credentials missing.

    Args:
        ai_model: Model identifier (e.g., "gpt-5-mini", "groq-llama")

    Returns:
        Provider name (openai, groq, gemini, openrouter, etc.)

    Raises:
        RuntimeError: If required API key is missing for the detected provider
    """
    provider = detect_provider(ai_model)

    logger.info(f"🔌 Model: {ai_model}")
    logger.info(f"📡 Provider detected: {provider.upper()}")

    # Validate required API keys are available
    if provider == "openai":
        if not os.getenv("OPENAI_API_KEY"):
            raise RuntimeError(
                "❌ OPENAI_API_KEY not found in environment. "
                "Please set it in .env file before running."
            )
        logger.info("✅ OPENAI_API_KEY found")

    elif provider == "groq":
        if not os.getenv("GROQ_API_KEY"):
            raise RuntimeError(
                "❌ GROQ_API_KEY not found in environment. "
                "Please set it in .env file before running."
            )
        logger.info("✅ GROQ_API_KEY found")

    elif provider == "gemini":
        if not os.getenv("GEMINI_API_KEY"):
            raise RuntimeError(
                "❌ GEMINI_API_KEY not found in environment. "
                "Please set it in .env file before running."
            )
        logger.info("✅ GEMINI_API_KEY found")

    else:  # openrouter
        if not os.getenv("OPENROUTER_API_KEY"):
            raise RuntimeError(
                "❌ OPENROUTER_API_KEY not found in environment. "
                "Please set it in .env file before running."
            )
        logger.info("✅ OPENROUTER_API_KEY found")

    logger.info(f"🚀 Client initialized for provider: {provider.upper()}")
    return provider


# ============================================================================
# PDF EXTRACTION
# ============================================================================


def extract_pdf_text_with_pages(pdf_path: Path, max_pages: int | None = None) -> str:
    """Extrae texto del PDF e inserta marcadores de página (1-indexed)."""
    reader = PdfReader(str(pdf_path))
    parts: list[str] = []
    num_pages = len(reader.pages)
    limit = min(num_pages, max_pages) if max_pages else num_pages

    for i in range(limit):
        page = reader.pages[i]
        text = page.extract_text() or ""
        # Limpieza mínima (sin destruir)
        text = text.replace("\x00", " ").strip()
        parts.append(PAGE_MARKER_FMT.format(page_num=i + 1) + text)

    return "\n".join(parts).strip()


def load_pdf_list(list_path: Path, input_dir: Path) -> list[Path]:
    """Carga una lista de PDFs desde un .txt (uno por línea)."""
    if not list_path.exists():
        raise RuntimeError(f"Archivo de lista no encontrado: {list_path}")

    entries = []
    for line in list_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        entries.append(line)

    if not entries:
        raise RuntimeError(f"Lista vacía en: {list_path}")

    has_relative = any(not Path(item).is_absolute() for item in entries)
    if has_relative and (not input_dir.exists() or not input_dir.is_dir()):
        raise RuntimeError(f"Directorio de entrada no válido para rutas relativas: {input_dir}")

    pdfs: list[Path] = []
    missing: list[str] = []
    seen: set[Path] = set()

    for item in entries:
        candidate = Path(item)
        if not candidate.is_absolute():
            candidate = input_dir / candidate
        if candidate.suffix.lower() != ".pdf":
            with_pdf = candidate.with_suffix(".pdf")
            if with_pdf.exists():
                candidate = with_pdf
        if candidate.exists() and candidate.is_file():
            resolved = candidate.resolve()
            if resolved not in seen:
                pdfs.append(resolved)
                seen.add(resolved)
        else:
            missing.append(item)

    if missing:
        logger.warning(
            "⚠️ PDFs no encontrados en lista (%s): %s",
            len(missing),
            ", ".join(missing[:5]) + ("..." if len(missing) > 5 else ""),
        )

    if not pdfs:
        raise RuntimeError(f"No se encontraron PDFs válidos en: {list_path}")

    return pdfs


def find_latest_jsonl(out_dir: Path, jsonl_name: str) -> Path | None:
    """Devuelve el JSONL con timestamp más reciente de una ejecución anterior.

    Cada ejecución escribe `analisis_DDMMYYYY_HHMMSS.jsonl`, así que para reanudar
    hay que localizar el último. Sin esto, `--skip-existing` comprobaba la
    existencia del fichero recién nombrado (que nunca existe) y reprocesaba todo.
    """
    if not out_dir.is_dir():
        return None

    stem = Path(jsonl_name).stem
    candidates = [p for p in out_dir.glob(f"{stem}_*.jsonl") if p.is_file()]
    if not candidates:
        return None

    return max(candidates, key=lambda p: p.stat().st_mtime)


def load_key_sentencias() -> set:
    """Carga el conjunto de nombres de sentencias clave desde archivo.

    Las sentencias clave se procesan con modelo premium (SENTENCIA_CLAVE_MODEL).
    Retorna set vacío si el archivo no existe.
    """
    if not KEY_SENTENCIAS_FILE.exists():
        logger.warning(f"⚠️ Archivo de sentencias clave no encontrado: {KEY_SENTENCIAS_FILE}")
        return set()

    key_sentencias = set()
    for line in KEY_SENTENCIAS_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#"):
            key_sentencias.add(line)

    logger.info(
        f"📌 Cargadas {len(key_sentencias)} sentencias clave (modelo: {SENTENCIA_CLAVE_MODEL})"
    )
    return key_sentencias


# ============================================================================
# SCHEMA CLEANING AND VALIDATION
# ============================================================================


def clean_schema(obj: dict[str, Any], filename: str) -> dict[str, Any]:
    """Limpia el schema eliminando keys no permitidas y normalizando enums."""

    # 1. Eliminar keys no permitidas
    if ENFORCE_ALLOWED_KEYS:
        keys_to_remove = [k for k in obj.keys() if k not in ALLOWED_KEYS]
        for key in keys_to_remove:
            logger.debug(f"🧹 {filename}: Eliminando key no permitida: {key}")
            del obj[key]

    # 2. Normalizar criterios (eliminar paréntesis y modificadores)
    for field in ["Criterios_residencia_detectados", "Criterio_decisivo"]:
        if field in obj and isinstance(obj[field], list):
            normalized = []
            for crit in obj[field]:
                if isinstance(crit, str):
                    # Extraer solo el CRIT_* base (antes de cualquier paréntesis o espacio)
                    base_crit = crit.split("(")[0].split(" ")[0].strip()
                    if base_crit in VALID_CRITERIOS:
                        if base_crit not in normalized:
                            normalized.append(base_crit)
                    elif base_crit.startswith("CRIT_"):
                        # Mapear a CRIT_OTRO si no es válido
                        if "CRIT_OTRO" not in normalized:
                            normalized.append("CRIT_OTRO")
            obj[field] = normalized

    # 3. Normalizar categorías de prueba en Pruebas_AEAT y Pruebas_contribuyente
    for field in ["Pruebas_AEAT", "Pruebas_contribuyente", "Pruebas_rechazadas_clave"]:
        if field in obj and isinstance(obj[field], list):
            for prueba in obj[field]:
                if isinstance(prueba, dict) and "categoria" in prueba:
                    cat = prueba["categoria"]
                    if isinstance(cat, str):
                        # Limpiar espacios, typos comunes
                        cat_clean = cat.replace(" Y_", "_Y_").replace("_ ", "_").strip()
                        if cat_clean not in VALID_CATEGORIAS_PRUEBA:
                            # Intentar mapear categorías cercanas
                            if "VINCULOS" in cat_clean and "ADMIN" in cat_clean:
                                cat_clean = "VINCULOS_ADMINISTRATIVOS_EN_ESPANA"
                            elif "FAMILIA" in cat_clean or "PERSONAL" in cat_clean:
                                cat_clean = "FAMILIA_Y_ENTORNO_PERSONAL"
                            elif "ACTIVIDAD" in cat_clean or "ECONOMICA" in cat_clean:
                                cat_clean = "ACTIVIDAD_ECONOMICA_Y_GESTION"
                            elif "DOCUM" in cat_clean and "FISCAL" in cat_clean:
                                cat_clean = "DOCUMENTACION_FISCAL_EXTRANJERA"
                            else:
                                # Si no se puede mapear, usar OTROS y mover detalle a subcategoria
                                if "subcategoria" not in prueba or not prueba["subcategoria"]:
                                    prueba["subcategoria"] = cat[:30]  # Máx 30 chars
                                cat_clean = "OTROS"
                        prueba["categoria"] = cat_clean

    # 4. Normalizar tiebreaker_paso_decisivo
    if "tiebreaker_paso_decisivo" in obj:
        tb = obj["tiebreaker_paso_decisivo"]
        if isinstance(tb, str) and tb not in VALID_TIEBREAKER_PASOS:
            # Intentar mapear
            tb_upper = tb.upper().replace(" ", "_")
            if "VIVIENDA" in tb_upper:
                obj["tiebreaker_paso_decisivo"] = "VIVIENDA_PERMANENTE"
            elif "VITAL" in tb_upper:
                obj["tiebreaker_paso_decisivo"] = "CENTRO_INTERESES_VITALES"
            elif "MORADA" in tb_upper:
                obj["tiebreaker_paso_decisivo"] = "MORADA_HABITUAL"
            elif "NACIONAL" in tb_upper:
                obj["tiebreaker_paso_decisivo"] = "NACIONALIDAD"
            elif "ACUERDO" in tb_upper or "MUTUO" in tb_upper:
                obj["tiebreaker_paso_decisivo"] = "ACUERDO_MUTUO"
            else:
                obj["tiebreaker_paso_decisivo"] = "NO_CONSTA"

    # 5. Normalizar resultado_final
    if "resultado_final" in obj:
        rf = obj["resultado_final"]
        if isinstance(rf, str):
            rf_norm = rf.upper().replace(" ", "_")
            # Mapear variantes comunes
            if "CONTRIBUYENTE" in rf_norm or "GANA" in rf_norm and "AEAT" not in rf_norm:
                obj["resultado_final"] = "GANA_CONTRIBUYENTE"
            elif "AEAT" in rf_norm:
                obj["resultado_final"] = "GANA_AEAT"
            elif "PARCIAL" in rf_norm:
                obj["resultado_final"] = "PARCIAL"
            elif "RETROAC" in rf_norm:
                obj["resultado_final"] = "RETROACCION"
            elif "INADMIS" in rf_norm:
                obj["resultado_final"] = "INADMISION"
            elif "FUERA" in rf_norm or "ALCANCE" in rf_norm:
                obj["resultado_final"] = "FUERA_DE_ALCANCE"
            elif rf_norm not in VALID_RESULTADO_FINAL:
                obj["resultado_final"] = "OTROS"

    return obj


def ensure_required_keys(obj: dict[str, Any], filename: str) -> dict[str, Any]:
    """Asegura que haya claves mínimas para no romper el pipeline."""

    def set_default(path: str, default: Any) -> None:
        if path not in obj:
            obj[path] = default

    obj["archivo"] = filename

    for key, default_value in REQUIRED_FIELDS.items():
        if key == "archivo":
            continue

        if isinstance(default_value, dict):
            set_default(key, dict(default_value))
        elif isinstance(default_value, list):
            set_default(key, [])
        else:
            set_default(key, default_value)

    # Aplicar limpieza de schema
    obj = clean_schema(obj, filename)

    return obj


# ============================================================================
# CSV FLATTENING
# ============================================================================


def flatten_for_csv(obj: dict[str, Any]) -> dict[str, Any]:
    """Aplana a columnas 'amigables CSV'."""

    def jdump(x: Any) -> str:
        return json.dumps(x, ensure_ascii=False)

    row: dict[str, Any] = {}

    # Identificación y filtrado
    row["archivo"] = obj.get("archivo", DEFAULT_MISSING_VALUE)
    row["es_caso_residencia_irpf"] = obj.get("es_caso_residencia_irpf", DEFAULT_MISSING_VALUE)
    row["fuera_de_alcance"] = obj.get("motivo_fuera_de_alcance", DEFAULT_MISSING_VALUE)

    ids = obj.get("identificadores", {}) or {}
    row["ROJ"] = ids.get("ROJ", DEFAULT_MISSING_VALUE)
    row["ECLI"] = ids.get("ECLI", DEFAULT_MISSING_VALUE)

    row["organo"] = obj.get("organo", DEFAULT_MISSING_VALUE)
    row["fecha_resolucion"] = obj.get("fecha_resolucion", DEFAULT_MISSING_VALUE)
    row["ejercicios_afectados"] = obj.get("ejercicios_afectados", DEFAULT_MISSING_VALUE)

    # Residencia y CDI
    row["pais_alegado_residencia_pf"] = obj.get("pais_alegado_residencia_pf", DEFAULT_MISSING_VALUE)
    row["pais_CDI_aplicado"] = obj.get("pais_CDI_aplicado", DEFAULT_MISSING_VALUE)
    row["se_invoca_CDI"] = obj.get("se_invoca_CDI", DEFAULT_MISSING_VALUE)
    row["tiebreaker_paso_decisivo"] = obj.get("tiebreaker_paso_decisivo", DEFAULT_MISSING_VALUE)

    # Criterios
    row["criterios_detectados"] = jdump(obj.get("Criterios_residencia_detectados", []))
    row["criterio_decisivo"] = jdump(obj.get("Criterio_decisivo", []))
    row["resumen_criterios"] = obj.get("resumen_criterios", DEFAULT_MISSING_VALUE)

    # Razonamiento judicial (nuevos campos)
    row["doctrina_citada"] = jdump(obj.get("doctrina_citada", []))
    row["carga_prueba"] = jdump(obj.get("carga_prueba", {}))
    row["razonamiento_residencia"] = obj.get("razonamiento_residencia", DEFAULT_MISSING_VALUE)

    # Pruebas detalladas
    row["pruebas_aeat"] = jdump(obj.get("Pruebas_AEAT", []))
    row["pruebas_contribuyente"] = jdump(obj.get("Pruebas_contribuyente", []))

    # Agregados para análisis (admitidas/rechazadas por parte)
    row["categorias_admitidas_aeat"] = jdump(obj.get("categorias_admitidas_aeat", []))
    row["categorias_rechazadas_aeat"] = jdump(obj.get("categorias_rechazadas_aeat", []))
    row["categorias_admitidas_contribuyente"] = jdump(
        obj.get("categorias_admitidas_contribuyente", [])
    )
    row["categorias_rechazadas_contribuyente"] = jdump(
        obj.get("categorias_rechazadas_contribuyente", [])
    )

    # Pruebas clave
    row["pruebas_rechazadas_clave"] = jdump(obj.get("Pruebas_rechazadas_clave", []))
    row["bala_de_plata"] = jdump(obj.get("Prueba_o_bala_de_plata", {}))

    # Resultado
    row["resultado_final"] = obj.get("resultado_final", DEFAULT_MISSING_VALUE)

    # Metadata
    row["frases_clave"] = jdump(obj.get("frases_clave", []))
    row["confianza_extraccion"] = obj.get("confianza_extraccion", DEFAULT_MISSING_VALUE)
    row["observaciones"] = obj.get("observaciones", DEFAULT_MISSING_VALUE)

    # Ejecución y costes
    row["tiempo_ejecucion"] = obj.get("tiempo_ejecucion", DEFAULT_MISSING_VALUE)
    row["costo_usd"] = obj.get("costo_usd", 0.0)

    # Incluir automáticamente cualquier campo top-level no mapeado
    for key, value in obj.items():
        if key in row:
            continue
        if isinstance(value, (dict, list)):
            row[key] = jdump(value)
        else:
            row[key] = value

    return row


# ============================================================================
# EXCEL EXPORT (DOS PESTAÑAS: SENTENCIAS + PRUEBAS)
# ============================================================================


def flatten_sentencia_for_excel(obj: dict[str, Any]) -> dict[str, Any]:
    """Aplana datos de sentencia para pestaña 'Sentencias' (sin pruebas detalladas)."""
    row: dict[str, Any] = {}

    # Identificación
    row["archivo"] = obj.get("archivo", DEFAULT_MISSING_VALUE)
    ids = obj.get("identificadores", {}) or {}
    row["ROJ"] = ids.get("ROJ", DEFAULT_MISSING_VALUE)
    row["ECLI"] = ids.get("ECLI", DEFAULT_MISSING_VALUE)
    row["organo"] = obj.get("organo", DEFAULT_MISSING_VALUE)
    row["fecha_resolucion"] = obj.get("fecha_resolucion", DEFAULT_MISSING_VALUE)

    # Filtrado
    row["es_caso_residencia_irpf"] = obj.get("es_caso_residencia_irpf", DEFAULT_MISSING_VALUE)
    row["motivo_fuera_de_alcance"] = obj.get("motivo_fuera_de_alcance", DEFAULT_MISSING_VALUE)
    row["ejercicios_afectados"] = obj.get("ejercicios_afectados", DEFAULT_MISSING_VALUE)

    # Residencia y CDI
    row["pais_alegado_residencia_pf"] = obj.get("pais_alegado_residencia_pf", DEFAULT_MISSING_VALUE)
    row["pais_CDI_aplicado"] = obj.get("pais_CDI_aplicado", DEFAULT_MISSING_VALUE)
    row["se_invoca_CDI"] = obj.get("se_invoca_CDI", DEFAULT_MISSING_VALUE)
    row["tiebreaker_paso_decisivo"] = obj.get("tiebreaker_paso_decisivo", DEFAULT_MISSING_VALUE)

    # Criterios
    criterios = obj.get("Criterios_residencia_detectados", [])
    row["criterios_detectados"] = ", ".join(criterios) if criterios else DEFAULT_MISSING_VALUE
    criterio_dec = obj.get("Criterio_decisivo", [])
    row["criterio_decisivo"] = ", ".join(criterio_dec) if criterio_dec else DEFAULT_MISSING_VALUE
    row["resumen_criterios"] = obj.get("resumen_criterios", DEFAULT_MISSING_VALUE)

    # Razonamiento judicial
    doctrina = obj.get("doctrina_citada", [])
    row["doctrina_citada"] = ", ".join(doctrina) if doctrina else DEFAULT_MISSING_VALUE

    carga = obj.get("carga_prueba", {}) or {}
    row["carga_prueba_quien"] = carga.get("quien_tenia_carga", DEFAULT_MISSING_VALUE)
    row["carga_prueba_cumplida"] = carga.get("cumplida", DEFAULT_MISSING_VALUE)
    row["carga_prueba_motivo"] = carga.get("motivo", DEFAULT_MISSING_VALUE)

    row["razonamiento_residencia"] = obj.get("razonamiento_residencia", DEFAULT_MISSING_VALUE)

    # Agregados de pruebas
    pruebas_aeat = obj.get("Pruebas_AEAT", []) or []
    pruebas_contrib = obj.get("Pruebas_contribuyente", []) or []

    row["total_pruebas_aeat"] = len(pruebas_aeat)
    row["total_pruebas_contribuyente"] = len(pruebas_contrib)
    row["pruebas_aeat_aceptadas"] = sum(1 for p in pruebas_aeat if p.get("aceptada") == "SI")
    row["pruebas_aeat_rechazadas"] = sum(1 for p in pruebas_aeat if p.get("aceptada") == "NO")
    row["pruebas_contrib_aceptadas"] = sum(1 for p in pruebas_contrib if p.get("aceptada") == "SI")
    row["pruebas_contrib_rechazadas"] = sum(1 for p in pruebas_contrib if p.get("aceptada") == "NO")

    # Categorías
    row["categorias_admitidas_aeat"] = ", ".join(obj.get("categorias_admitidas_aeat", []))
    row["categorias_rechazadas_aeat"] = ", ".join(obj.get("categorias_rechazadas_aeat", []))
    row["categorias_admitidas_contribuyente"] = ", ".join(
        obj.get("categorias_admitidas_contribuyente", [])
    )
    row["categorias_rechazadas_contribuyente"] = ", ".join(
        obj.get("categorias_rechazadas_contribuyente", [])
    )

    # Bala de plata
    bala = obj.get("Prueba_o_bala_de_plata", {}) or {}
    row["bala_plata_parte"] = bala.get("parte", DEFAULT_MISSING_VALUE)
    row["bala_plata_categoria"] = bala.get("categoria", DEFAULT_MISSING_VALUE)
    row["bala_plata_detalle"] = bala.get("detalle", DEFAULT_MISSING_VALUE)
    row["bala_plata_por_que"] = bala.get("por_que_decisiva", DEFAULT_MISSING_VALUE)

    # Resultado
    row["resultado_final"] = obj.get("resultado_final", DEFAULT_MISSING_VALUE)
    row["confianza_extraccion"] = obj.get("confianza_extraccion", DEFAULT_MISSING_VALUE)
    row["observaciones"] = obj.get("observaciones", DEFAULT_MISSING_VALUE)

    # Metadata
    row["tiempo_ejecucion"] = obj.get("tiempo_ejecucion", DEFAULT_MISSING_VALUE)
    row["costo_usd"] = obj.get("costo_usd", 0.0)

    return row


def expand_pruebas_for_excel(obj: dict[str, Any]) -> list[dict[str, Any]]:
    """Expande cada prueba a una fila para pestaña 'Pruebas'."""
    rows = []
    archivo = obj.get("archivo", "unknown")
    resultado = obj.get("resultado_final", DEFAULT_MISSING_VALUE)

    # Pruebas AEAT
    for prueba in obj.get("Pruebas_AEAT", []) or []:
        row = {
            "archivo": archivo,
            "resultado_sentencia": resultado,
            "parte": "AEAT",
            "categoria": prueba.get("categoria", DEFAULT_MISSING_VALUE),
            "subcategoria": prueba.get("subcategoria", DEFAULT_MISSING_VALUE),
            "detalle": prueba.get("detalle", DEFAULT_MISSING_VALUE),
            "objetivo_probatorio": prueba.get("objetivo_probatorio", DEFAULT_MISSING_VALUE),
            "criterio_atacado": prueba.get("criterio_atacado", DEFAULT_MISSING_VALUE),
            "tipo_prueba": prueba.get("tipo_prueba", DEFAULT_MISSING_VALUE),
            "origen": prueba.get("origen", DEFAULT_MISSING_VALUE),
            "aceptada": prueba.get("aceptada", DEFAULT_MISSING_VALUE),
            "peso": prueba.get("peso", 0),
            "motivo_valoracion": prueba.get("motivo_valoracion", DEFAULT_MISSING_VALUE),
            "contradiccion_con": prueba.get("contradiccion_con", DEFAULT_MISSING_VALUE),
            "cita_pagina": (prueba.get("cita") or {}).get("pagina", DEFAULT_MISSING_VALUE),
            "cita_texto": (prueba.get("cita") or {}).get("texto", DEFAULT_MISSING_VALUE),
        }
        rows.append(row)

    # Pruebas CONTRIBUYENTE
    for prueba in obj.get("Pruebas_contribuyente", []) or []:
        row = {
            "archivo": archivo,
            "resultado_sentencia": resultado,
            "parte": "CONTRIBUYENTE",
            "categoria": prueba.get("categoria", DEFAULT_MISSING_VALUE),
            "subcategoria": prueba.get("subcategoria", DEFAULT_MISSING_VALUE),
            "detalle": prueba.get("detalle", DEFAULT_MISSING_VALUE),
            "objetivo_probatorio": prueba.get("objetivo_probatorio", DEFAULT_MISSING_VALUE),
            "criterio_atacado": prueba.get("criterio_atacado", DEFAULT_MISSING_VALUE),
            "tipo_prueba": prueba.get("tipo_prueba", DEFAULT_MISSING_VALUE),
            "origen": prueba.get("origen", DEFAULT_MISSING_VALUE),
            "aceptada": prueba.get("aceptada", DEFAULT_MISSING_VALUE),
            "peso": prueba.get("peso", 0),
            "motivo_valoracion": prueba.get("motivo_valoracion", DEFAULT_MISSING_VALUE),
            "contradiccion_con": prueba.get("contradiccion_con", DEFAULT_MISSING_VALUE),
            "cita_pagina": (prueba.get("cita") or {}).get("pagina", DEFAULT_MISSING_VALUE),
            "cita_texto": (prueba.get("cita") or {}).get("texto", DEFAULT_MISSING_VALUE),
        }
        rows.append(row)

    return rows


def _sanitize_cell_value(value: Any) -> Any:
    """Sanitiza valores para Excel: elimina caracteres de control no permitidos."""
    if not isinstance(value, str):
        return value
    # Elimina caracteres de control (ASCII 0-31) excepto tab(9), newline(10), CR(13)
    # Excel no permite estos caracteres en las celdas
    return "".join(c for c in value if ord(c) >= 32 or c in "\t\n\r")


def generate_normalized_exports(
    jsonl_path: Path, output_dir: Path, timestamp: str, logger: logging.Logger
) -> None:
    """Genera CSVs normalizados y Excel con dos pestañas."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    sentencias_rows = []
    pruebas_rows = []

    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                sentencias_rows.append(flatten_sentencia_for_excel(obj))
                pruebas_rows.extend(expand_pruebas_for_excel(obj))
            except Exception:
                continue

    # Crear DataFrames
    df_sentencias = pd.DataFrame(sentencias_rows)
    df_pruebas = pd.DataFrame(pruebas_rows)

    # Rutas de salida
    sentencias_csv = output_dir / f"sentencias_{timestamp}.csv"
    pruebas_csv = output_dir / f"pruebas_{timestamp}.csv"
    xlsx_path = output_dir / f"analisis_{timestamp}.xlsx"

    # 1. Guardar CSVs
    df_sentencias.to_csv(sentencias_csv, index=False, encoding="utf-8")
    logger.info(f"📄 CSV Sentencias: {sentencias_csv} ({len(df_sentencias)} filas)")

    df_pruebas.to_csv(pruebas_csv, index=False, encoding="utf-8")
    logger.info(f"📄 CSV Pruebas: {pruebas_csv} ({len(df_pruebas)} filas)")

    # 2. Crear Excel con dos hojas
    wb = Workbook()

    # Hoja 1: Sentencias
    ws_sentencias = wb.active
    ws_sentencias.title = "Sentencias"
    for r_idx, row in enumerate(dataframe_to_rows(df_sentencias, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            clean_value = _sanitize_cell_value(value)
            ws_sentencias.cell(row=r_idx, column=c_idx, value=clean_value)

    # Hoja 2: Pruebas
    ws_pruebas = wb.create_sheet("Pruebas")
    for r_idx, row in enumerate(dataframe_to_rows(df_pruebas, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            clean_value = _sanitize_cell_value(value)
            ws_pruebas.cell(row=r_idx, column=c_idx, value=clean_value)

    wb.save(xlsx_path)
    logger.info(f"📊 Excel: {xlsx_path}")
    logger.info(f"   - Pestaña 'Sentencias': {len(df_sentencias)} filas")
    logger.info(f"   - Pestaña 'Pruebas': {len(df_pruebas)} filas")


# ============================================================================
# MAIN ASYNC PROCESSING LOOP
# ============================================================================


async def process_pdf_async(
    pdf_path: Path,
    ai_model: str,
    max_pages: int | None,
    reasoning_effort: str | None = None,
) -> dict[str, Any]:
    """Procesa un PDF usando gpt_request."""
    fname = pdf_path.name

    try:
        pdf_text = extract_pdf_text_with_pages(pdf_path, max_pages=max_pages)

        if not pdf_text.strip():
            logger.warning(f"⚠️ {fname}: No se pudo extraer texto")
            return ensure_required_keys(
                {
                    "archivo": fname,
                    "observaciones": "PDF sin texto extraíble",
                    "confianza_extraccion": "BAJA",
                },
                fname,
            )

        # Usar gpt_request si está disponible
        if USE_GPT_REQUEST:
            effort_log = reasoning_effort if reasoning_effort else "default"
            logger.info(
                f"📨 Procesando {fname} con gpt_request ({ai_model}, reasoning_effort={effort_log})"
            )
            result = await gpt_request_for_sentencia(
                ai_model=ai_model,
                system_prompt=SYSTEM_PROMPT,
                pdf_text=pdf_text,
                logger=logger,
                temperature=0,
                response_format="json_object",
                reasoning_effort=reasoning_effort if "gpt-5" in ai_model else None,
            )

            if "error" in result:
                logger.error(f"❌ {fname}: Error en gpt_request: {result.get('error')}")
                return ensure_required_keys(
                    {
                        "archivo": fname,
                        "observaciones": f"ERROR_GPT_REQUEST: {result.get('error')}",
                        "confianza_extraccion": "BAJA",
                    },
                    fname,
                )

            # Extraer información de tiempo y coste
            tiempo_ejecucion = result.pop("tiempo_ejecucion", "NO CONSTA")
            cost_usd = result.pop("cost_usd", 0)

            # Eliminar tokens (no se guardan)
            result.pop("tokens_in", None)
            result.pop("tokens_out", None)

            # Eliminar otros campos de metadata
            obj = {k: v for k, v in result.items() if k not in ["error"]}

            # Agregar metadata de ejecución y coste directamente en el objeto
            obj["tiempo_ejecucion"] = tiempo_ejecucion
            obj["costo_usd"] = cost_usd
        else:
            logger.warning("⚠️ gpt_request no disponible, usando fallback")
            return ensure_required_keys(
                {
                    "archivo": fname,
                    "observaciones": "ERROR: gpt_request no disponible",
                    "confianza_extraccion": "BAJA",
                },
                fname,
            )

        return ensure_required_keys(obj, fname)

    except Exception as e:
        logger.error(f"🚨 Error procesando {fname}: {e}")
        return ensure_required_keys(
            {
                "archivo": fname,
                "observaciones": f"ERROR_PROCESO: {str(e)}",
                "confianza_extraccion": "BAJA",
            },
            fname,
        )


async def main_async(
    in_dir: Path,
    out_dir: Path,
    jsonl_path: Path,
    csv_path: Path,
    ai_model: str,
    max_pages: int | None,
    max_files: int | None,
    skip_existing: bool,
    pdf_list: list[Path] | None = None,
    reasoning_effort: str | None = None,
    timestamp: str = "",
) -> None:
    """Bucle principal de procesamiento en batches paralelos.

    Args:
        max_files: Máximo número de PDFs a procesar (0 o None = sin límite)
        reasoning_effort: Reasoning effort level (low, medium, high)
    """

    if pdf_list is not None:
        pdf_files = list(pdf_list)
    else:
        if not in_dir.exists() or not in_dir.is_dir():
            raise RuntimeError(f"Carpeta de entrada no válida: {in_dir}")
        # Buscar PDFs recursivamente
        pdf_files = sorted([p for p in in_dir.glob("**/*.pdf") if p.is_file()])
        if not pdf_files:
            raise RuntimeError(f"No encontré PDFs en: {in_dir}")

    # Limitar PDFs si se especifica max_files
    if max_files and max_files > 0:
        pdf_files = pdf_files[:max_files]

    logger.info(f"📁 Encontrados {len(pdf_files)} PDFs para procesar")

    # Cargar ya procesados si skip-existing
    processed = set()
    if skip_existing and jsonl_path.exists():
        with jsonl_path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                    processed.add(obj.get("archivo"))
                except Exception:
                    continue

    # Filtrar PDFs ya procesados
    pdfs_to_process = [p for p in pdf_files if not (skip_existing and p.name in processed)]
    logger.info(
        f"📄 Procesarán {len(pdfs_to_process)} PDFs (saltando {len(pdf_files) - len(pdfs_to_process)} ya procesados)"
    )

    # Cargar sentencias clave (modelo premium)
    key_sentencias = load_key_sentencias()
    key_in_batch = [p.name for p in pdfs_to_process if p.name in key_sentencias]
    if key_in_batch:
        logger.info(
            f"   🔑 {len(key_in_batch)} sentencias clave en cola (usarán {SENTENCIA_CLAVE_MODEL})"
        )

    logger.info(f"⚡ Modo paralelo: {BATCH_SIZE} PDFs por batch\n")

    jsonl_mode = "a" if jsonl_path.exists() else "w"
    total_cost = 0.0
    batch_costs = {}

    with jsonl_path.open(jsonl_mode, encoding="utf-8") as jf:
        total_pdfs = len(pdfs_to_process)
        for batch_idx in range(0, total_pdfs, BATCH_SIZE):
            batch = pdfs_to_process[batch_idx : batch_idx + BATCH_SIZE]
            batch_num = batch_idx // BATCH_SIZE + 1
            total_batches = (total_pdfs + BATCH_SIZE - 1) // BATCH_SIZE

            # Mostrar inicio del batch
            pdf_names = ", ".join([p.name for p in batch[:3]])
            if len(batch) > 3:
                pdf_names += f", ... +{len(batch) - 3} más"
            logger.info(
                f"⚡ Batch {batch_num}/{total_batches} | Procesando {len(batch)} PDFs en paralelo"
            )
            logger.info(f"   📁 {pdf_names}")

            # Procesar todos los PDFs en el batch de forma concurrente
            # Usar modelo premium para sentencias clave
            def get_model_for_pdf(pdf_path: Path) -> str:
                if pdf_path.name in key_sentencias:
                    return SENTENCIA_CLAVE_MODEL
                return ai_model

            results = await asyncio.gather(
                *[
                    process_pdf_async(
                        pdf_path, get_model_for_pdf(pdf_path), max_pages, reasoning_effort
                    )
                    for pdf_path in batch
                ]
            )

            # Guardar resultados en JSONL y acumular costes
            batch_cost = 0.0
            for obj in results:
                # Extraer coste (ya está en el objeto como costo_usd)
                cost_usd = obj.get("costo_usd", 0.0)
                batch_cost += cost_usd
                total_cost += cost_usd

                jf.write(json.dumps(obj, ensure_ascii=False) + "\n")

                # Log por PDF
                pdf_name = obj.get("archivo", "unknown")
                if cost_usd > 0:
                    logger.debug(f"   💰 {pdf_name}: ${cost_usd:.4f}")

            jf.flush()

            # Mostrar progreso detallado con coste
            processed_count = batch_idx + len(batch)
            percentage = (processed_count / total_pdfs) * 100
            logger.info(
                f"✅ Progreso: {processed_count}/{total_pdfs} PDFs completados ({percentage:.1f}%) | Batch: ${batch_cost:.2f}\n"
            )
            batch_costs[batch_num] = batch_cost

    # Convertir JSONL -> CSV
    logger.info("🔄 Convirtiendo JSONL a CSV...")
    rows: list[dict[str, Any]] = []
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                rows.append(flatten_for_csv(obj))
            except Exception:
                continue

    df = pd.DataFrame(rows)

    # Reordenar columnas
    cols = [c for c in CSV_COLUMN_ORDER if c in df.columns] + [
        c for c in df.columns if c not in CSV_COLUMN_ORDER
    ]
    df = df[cols]

    df.to_csv(csv_path, index=False, encoding="utf-8")

    # Generar exports normalizados (CSVs + Excel con pestañas)
    if timestamp:
        generate_normalized_exports(jsonl_path, out_dir, timestamp, logger)

    # Mostrar resumen final con costes
    logger.info(f"\n{'=' * 60}")
    logger.info("✅ PROCESAMIENTO COMPLETADO")
    logger.info(f"{'=' * 60}")
    logger.info(f"📄 JSONL: {jsonl_path}")
    logger.info(f"📊 CSV:   {csv_path}")
    logger.info(f"📈 Filas: {len(df)}")
    logger.info("\n💰 COSTES DE API:")
    logger.info(f"   Total: ${total_cost:.2f} USD")
    logger.info(f"   PDFs procesados: {len(pdfs_to_process)}")
    if len(pdfs_to_process) > 0:
        logger.info(f"   Coste promedio: ${total_cost / len(pdfs_to_process):.4f} USD por PDF")

    if batch_costs:
        logger.info("\n📊 Desglose por batch:")
        for batch_num in sorted(batch_costs.keys()):
            logger.info(f"   Batch {batch_num}: ${batch_costs[batch_num]:.2f}")

    logger.info(f"{'=' * 60}\n")


def main() -> None:
    """Punto de entrada principal."""
    load_dotenv()

    parser = argparse.ArgumentParser(description=SCRIPT_DESCRIPTION)
    parser.add_argument("--input", default=str(DEFAULT_INPUT_DIR), help=ARGUMENT_HELP["input"])
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_DIR), help=ARGUMENT_HELP["output"])
    parser.add_argument("--model", default=DEFAULT_MODEL, help=ARGUMENT_HELP["model"])
    parser.add_argument("--pdf-list", help=ARGUMENT_HELP["pdf_list"])
    parser.add_argument(
        "--max-files", type=int, default=DEFAULT_MAX_FILES, help=ARGUMENT_HELP["max_files"]
    )
    parser.add_argument(
        "--jsonl-name", default=DEFAULT_JSONL_NAME, help=ARGUMENT_HELP["jsonl_name"]
    )
    parser.add_argument("--csv-name", default=DEFAULT_CSV_NAME, help=ARGUMENT_HELP["csv_name"])
    parser.add_argument("--skip-existing", action="store_true", help=ARGUMENT_HELP["skip_existing"])
    parser.add_argument("--resume-from", help=ARGUMENT_HELP["resume_from"])
    parser.add_argument(
        "--reasoning-effort",
        default=REASONING_EFFORT,
        choices=["low", "medium", "high"],
        help="Reasoning effort level for GPT-5 models (default: medium)",
    )
    args = parser.parse_args()

    in_dir = Path(args.input).expanduser().resolve()
    out_dir = Path(args.output).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    # Agregar timestamp a los nombres de salida (DDMMYYYY_HHMMSS)
    timestamp = get_timestamp_suffix()
    jsonl_name = args.jsonl_name.replace(".jsonl", f"_{timestamp}.jsonl")
    csv_name = args.csv_name.replace(".csv", f"_{timestamp}.csv")

    jsonl_path = out_dir / jsonl_name
    csv_path = out_dir / csv_name

    # Reanudación: hay que apuntar el JSONL a la ejecución anterior, no al fichero
    # con timestamp nuevo. El CSV y los exports sí se regeneran completos desde el
    # JSONL resultante, así que se quedan con el timestamp de esta ejecución.
    skip_existing = args.skip_existing or bool(args.resume_from)
    if args.resume_from:
        resume_path: Path | None = Path(args.resume_from).expanduser().resolve()
        if resume_path is not None and not resume_path.is_file():
            raise RuntimeError(f"El JSONL indicado en --resume-from no existe: {resume_path}")
    elif args.skip_existing:
        resume_path = find_latest_jsonl(out_dir, args.jsonl_name)
        if resume_path is None:
            logger.warning(
                "⚠️ --skip-existing sin JSONL previo en %s: se procesarán todos los PDFs",
                out_dir,
            )
    else:
        resume_path = None

    if resume_path is not None:
        jsonl_path = resume_path
        logger.info(f"♻️ Reanudando sobre {jsonl_path.name} (se le añadirán los nuevos resultados)")

    max_pages = None
    max_files = args.max_files if args.max_files and args.max_files > 0 else None

    logger.info("🚀 Iniciando procesamiento de sentencias")
    logger.info(f"   📁 Entrada: {in_dir}")
    logger.info(f"   📤 Salida: {out_dir}")
    logger.info(f"   🤖 Modelo: {args.model}")
    logger.info(f"   ⚙️ Reasoning Effort: {args.reasoning_effort}")
    logger.info(f"   ⏰ Timestamp: {timestamp}")
    logger.info(f"   📋 JSONL: {jsonl_name}")
    logger.info(f"   📊 CSV: {csv_name}")
    if max_files:
        logger.info(f"   📄 Límite de archivos: {max_files}")

    # Initialize client for selected model (fails fast if API key missing)
    try:
        provider = initialize_client(args.model)
        logger.info(f"   ✅ Provider: {provider.upper()}")
    except RuntimeError as e:
        logger.error(str(e))
        raise

    pdf_list = None
    if args.pdf_list:
        list_path = Path(args.pdf_list).expanduser().resolve()
        pdf_list = load_pdf_list(list_path, in_dir)
        logger.info(f"   📄 Lista PDF: {list_path} ({len(pdf_list)} archivos)")

    # Ejecutar bucle async
    asyncio.run(
        main_async(
            in_dir=in_dir,
            out_dir=out_dir,
            jsonl_path=jsonl_path,
            csv_path=csv_path,
            ai_model=args.model,
            max_pages=max_pages,
            max_files=max_files,
            skip_existing=skip_existing,
            pdf_list=pdf_list,
            reasoning_effort=args.reasoning_effort,
            timestamp=timestamp,
        )
    )


if __name__ == "__main__":
    main()
